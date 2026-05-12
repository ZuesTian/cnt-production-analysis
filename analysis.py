# -*- coding: utf-8 -*-
"""
碳纳米管生产数据分析脚本

功能：
1. 炉子级统计：每个炉子的每次反应周期明细、月度平均、每月前/后20%筛选。
2. 每日汇总：总产量、总反应时间、总故障时间、总空烧时间、总降清时间，并生成趋势图。
3. 每月汇总：同每日口径按月汇总，并生成趋势图。

说明：
原始表只有“停机清理空烧”一列，没有单独的“降清时间”列。本脚本按当前数据口径，
将该列同时作为“空烧时间”和“降清时间”输出。后续如果源表拆分字段，只需调整
SOURCE_COLUMNS 中的映射。
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import io
import json
import math
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

import matplotlib
import yaml

matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.font_manager as font_manager
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("output")
CONFIG_FILE = Path("config.yaml")
EXCEL_BASE_DATE = datetime(1899, 12, 30)
SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

DEFAULT_SOURCE_COLUMNS = {
    "date": "日期",
    "team": "班组",
    "furnace": "炉号",
    "reaction_time": "生产时间",
    "fault_time": "设备故障影响时间",
    "clean_empty_burn_time": "停机清理空烧",
    "output": "产量",
    "source_yield": "小时产能",
}

DEFAULT_PRODUCTION_LINES = {
    "L3": {"patterns": ["L3", "E", "F", "G", "H", "B"]},
    "11A": {"patterns": ["11A"]},
}

DEFAULT_ALERT_THRESHOLDS = {
    "max_fault_hours_per_day": 24,
    "fault_warning_hours_per_day": 8,
    "fault_critical_hours_per_day": 12,
    "consecutive_fault_days": 2,
    "monthly_fault_hours_warning": 24,
    "min_yield_rate": 50,
    "anomaly_sigma": 2.0,
}


def load_config(config_file: Path = CONFIG_FILE) -> dict:
    if not config_file.exists():
        return {}
    with config_file.open("r", encoding="utf-8") as stream:
        config = yaml.safe_load(stream) or {}
    if not isinstance(config, dict):
        raise ValueError(f"配置文件格式错误：{config_file}")
    return config


CONFIG = load_config()
SOURCE_COLUMNS = {**DEFAULT_SOURCE_COLUMNS, **CONFIG.get("source_columns", {})}
PRODUCTION_LINES = {**DEFAULT_PRODUCTION_LINES, **CONFIG.get("production_lines", {})}
RANKING_TOP_PERCENT = float(CONFIG.get("ranking", {}).get("top_percent", 0.2))
ALERT_THRESHOLDS = {**DEFAULT_ALERT_THRESHOLDS, **CONFIG.get("alert_thresholds", {})}

CYCLE_METRICS = ["反应时间", "空烧时间", "降清时间", "故障时间", "产率"]
SUMMARY_METRICS = ["总产量", "总反应时间", "总故障时间", "总空烧时间", "总降清时间", "平均产率"]
TIME_SUMMARY_METRICS = ["总反应时间", "总故障时间", "总空烧时间", "总降清时间"]
REQUIRED_SOURCE_COLUMNS = list(SOURCE_COLUMNS.values())
PREFERRED_CHINESE_FONTS = [
    "Microsoft YaHei",
    "SimHei",
    "Noto Sans CJK SC",
    "Noto Sans CJK JP",
    "Source Han Sans SC",
    "PingFang SC",
    "Hiragino Sans GB",
    "WenQuanYi Micro Hei",
    "Arial Unicode MS",
]

CHART_SURFACE = "#FFFFFF"
CHART_TEXT = "#18232F"
CHART_MUTED = "#667085"
CHART_GRID = "#E5EBF0"
CHART_BORDER = "#CBD5E1"
CHART_COLORS = {
    "production": "#177E76",
    "yield": "#B76E00",
    "reaction": "#2F66A3",
    "fault": "#C43C39",
    "empty_burn": "#D99224",
    "clean": "#7657A6",
    "neutral": "#64748B",
    "threshold": "#D92D20",
    "highlight": "#F97316",
    "highlight_alt": "#2563EB",
    "low": "#C43C39",
    "ok": "#177E76",
}


def configure_chinese_fonts() -> None:
    available_fonts = {font.name for font in font_manager.fontManager.ttflist}
    selected_fonts = [font for font in PREFERRED_CHINESE_FONTS if font in available_fonts]
    if not selected_fonts:
        print("警告：未检测到常见中文字体，图表中文可能显示异常。")
    plt.rcParams["font.sans-serif"] = selected_fonts + ["DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False


configure_chinese_fonts()


def _axes_list(axes) -> list:
    return list(np.ravel(np.asarray(axes, dtype=object)))


def _style_axis(ax, grid_axis: str | None = "y") -> None:
    ax.set_facecolor(CHART_SURFACE)
    ax.tick_params(axis="both", colors=CHART_MUTED, labelsize=9)
    ax.title.set_color(CHART_TEXT)
    ax.xaxis.label.set_color(CHART_MUTED)
    ax.yaxis.label.set_color(CHART_MUTED)
    if grid_axis:
        ax.grid(axis=grid_axis, color=CHART_GRID, linewidth=0.8, alpha=1.0)
    ax.set_axisbelow(True)
    for name, spine in ax.spines.items():
        spine.set_color(CHART_BORDER)
        spine.set_linewidth(0.8)
        if name in {"top", "right"}:
            spine.set_visible(False)
    with contextlib.suppress(Exception):
        ax.yaxis.set_major_locator(mticker.MaxNLocator(nbins=6))


def _style_figure(fig, axes, grid_axis: str | None = "y") -> None:
    fig.patch.set_facecolor(CHART_SURFACE)
    for ax in _axes_list(axes):
        _style_axis(ax, grid_axis=grid_axis)


def _style_legend(legend) -> None:
    if legend is None:
        return
    frame = legend.get_frame()
    frame.set_facecolor("#FFFFFF")
    frame.set_edgecolor(CHART_BORDER)
    frame.set_linewidth(0.8)
    for text in legend.get_texts():
        text.set_color(CHART_TEXT)


def _format_date_axis(ax, *, max_ticks: int = 10) -> None:
    locator = mdates.AutoDateLocator(minticks=4, maxticks=max_ticks)
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))


def _save_or_buffer_figure(fig, output_path: Path | None) -> Path | io.BytesIO:
    if output_path is None:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close(fig)
        buf.seek(0)
        return buf
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    return output_path


def configured_min_yield_rate() -> float | None:
    value = ALERT_THRESHOLDS.get("min_yield_rate")
    if value in (None, ""):
        return None
    try:
        rate = float(value)
    except (TypeError, ValueError):
        print(f"警告：min_yield_rate 配置无效，已忽略：{value}")
        return None
    if not np.isfinite(rate):
        return None
    return rate


@dataclass(frozen=True)
class AnalysisOutputs:
    furnace_workbook: Path | None = None
    furnace_chart: Path | None = None
    daily_workbook: Path | None = None
    daily_chart: Path | None = None
    monthly_workbook: Path | None = None
    monthly_chart: Path | None = None
    furnace_daily_workbook: Path | None = None
    furnace_daily_chart_dir: Path | None = None
    anomaly_workbook: Path | None = None


def find_input_file(path_arg: str | None = None) -> Path:
    """Resolve the source workbook path."""
    if path_arg:
        path = Path(path_arg)
        if not path.exists():
            raise FileNotFoundError(f"找不到输入文件：{path}")
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise ValueError(f"不支持的输入格式：{path.suffix}，请使用 .xlsx / .xls / .csv")
        return path

    candidates = [
        path
        for path in Path(".").iterdir()
        if not path.name.startswith("~$") and path.parent != OUTPUT_DIR
        and path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    ]
    if not candidates:
        raise FileNotFoundError("当前目录未找到 .xlsx / .xls / .csv 输入文件")

    exact = [path for path in candidates if "生产数据" in path.name]
    return sorted(exact or candidates, key=lambda p: p.name)[0]


def clean_number(value: object) -> float:
    """Convert common dirty Excel cell values to numeric values."""
    if pd.isna(value):
        return np.nan

    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "-"}:
        return np.nan

    text = (
        text.replace("..", ".")
        .replace(",", "")
        .replace("，", "")
        .replace("小时", "")
        .replace("kg", "")
        .strip()
    )
    return pd.to_numeric(text, errors="coerce")


def convert_excel_date(value: object) -> pd.Timestamp:
    """Convert Excel serial dates or regular date values to pandas Timestamp."""
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, (int, float, np.integer, np.floating)):
        return pd.Timestamp(EXCEL_BASE_DATE + timedelta(days=float(value))).normalize()

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return pd.NaT
    return pd.Timestamp(parsed).normalize()


def production_line_from_sheet(sheet_name: str) -> str:
    upper_name = str(sheet_name).strip().upper()
    configured = sorted(PRODUCTION_LINES.items(), key=lambda item: len(item[0]), reverse=True)
    for line, settings in configured:
        patterns = settings.get("patterns", []) if isinstance(settings, dict) else []
        for pattern in patterns:
            if upper_name.startswith(str(pattern).upper()):
                return line
    return "L3"


def production_line_from_furnace(furnace: object, fallback: str) -> str:
    upper_furnace = str(furnace).strip().upper()
    configured = sorted(
        PRODUCTION_LINES.items(),
        key=lambda item: max((len(str(pattern)) for pattern in item[1].get("patterns", [])), default=0),
        reverse=True,
    )
    for line, settings in configured:
        patterns = settings.get("patterns", []) if isinstance(settings, dict) else []
        for pattern in patterns:
            if upper_furnace.startswith(str(pattern).upper()):
                return line
    return fallback


def required_usecols(column: str) -> bool:
    return column in REQUIRED_SOURCE_COLUMNS


def read_csv_with_fallback(path: Path) -> pd.DataFrame:
    """Read CSV files with common Chinese encodings."""
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gbk", "gb18030"):
        try:
            return pd.read_csv(path, encoding=encoding, usecols=required_usecols)
        except UnicodeDecodeError as exc:
            last_error = exc

    if last_error:
        raise last_error
    return pd.read_csv(path, usecols=required_usecols)


def iter_source_tables(input_file: Path) -> Iterable[tuple[int, str, pd.DataFrame]]:
    """Yield source tables from Excel sheets or one CSV file."""
    if input_file.suffix.lower() == ".csv":
        yield 0, input_file.stem, read_csv_with_fallback(input_file)
        return

    excel = pd.ExcelFile(input_file)
    for sheet_index, sheet_name in enumerate(excel.sheet_names):
        yield sheet_index, sheet_name, pd.read_excel(input_file, sheet_name=sheet_name, usecols=required_usecols)


def missing_required_columns(df: pd.DataFrame) -> list[str]:
    return [col for col in REQUIRED_SOURCE_COLUMNS if col not in df.columns]


def format_missing_column_error(input_file: Path, skipped_tables: list[tuple[str, list[str]]]) -> str:
    details = "；".join(
        f"{sheet_name} 缺少 {', '.join(missing)}" for sheet_name, missing in skipped_tables
    )
    return (
        f"输入文件 {input_file} 未找到包含完整生产数据列的工作表。"
        f"必需列：{', '.join(REQUIRED_SOURCE_COLUMNS)}。"
        f"检查结果：{details}"
    )


def cache_key_for_input(input_file: Path) -> str:
    stat = input_file.stat()
    payload = {
        "path": str(input_file.resolve()),
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
        "source_columns": SOURCE_COLUMNS,
        "production_lines": PRODUCTION_LINES,
        "version": 3,
    }
    text = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:20]


def cache_path_for_input(input_file: Path) -> Path:
    return OUTPUT_DIR / ".cache" / f"cycles_{cache_key_for_input(input_file)}.pkl"


def load_cycles_cache(input_file: Path) -> pd.DataFrame | None:
    cache_path = cache_path_for_input(input_file)
    if not cache_path.exists():
        return None
    try:
        cycles = pd.read_pickle(cache_path)
    except Exception:
        return None
    print(f"  使用缓存数据：{cache_path}")
    return cycles


def save_cycles_cache(input_file: Path, cycles: pd.DataFrame) -> None:
    cache_path = cache_path_for_input(input_file)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cycles.to_pickle(cache_path)


def join_unique(values: pd.Series) -> str:
    unique_values = []
    for value in values.dropna():
        text = str(value).strip()
        if text and text.lower() != "nan" and text not in unique_values:
            unique_values.append(text)
    return "；".join(unique_values)


def join_line_numbers(values: pd.Series) -> str:
    numbers = []
    for value in values.dropna():
        try:
            numbers.append(int(value))
        except (TypeError, ValueError):
            continue
    return ",".join(str(number) for number in sorted(set(numbers)))


def sum_with_nan(values: pd.Series) -> float:
    return values.sum(min_count=1)


def weighted_yield(output: pd.Series, reaction_time: pd.Series) -> pd.Series:
    result = output / reaction_time
    result = result.mask(~np.isfinite(result), np.nan)
    return result


def add_weighted_average_yield(df: pd.DataFrame, output_col: str = "总产量", time_col: str = "总反应时间") -> pd.DataFrame:
    df["平均产率"] = weighted_yield(df[output_col], df[time_col]).round(2)
    return df


def build_reaction_cycles(raw: pd.DataFrame) -> pd.DataFrame:
    """Each cleaned row = one reaction cycle (one start-run-stop per shift)."""
    cycles = raw.copy()
    cycles["产率"] = cycles["产量"] / cycles["反应时间"]
    cycles.loc[~np.isfinite(cycles["产率"]), "产率"] = np.nan
    cycles["周期序号"] = cycles.groupby("炉号").cumcount() + 1

    ordered_cols = [
        "周期序号",
        "日期",
        "年月",
        "月份",
        "生产线",
        "班组",
        "炉号",
        "反应时间",
        "空烧时间",
        "降清时间",
        "故障时间",
        "产量",
        "产率",
        "源表小时产能",
        "来源工作表",
        "来源行号",
    ]
    return cycles[ordered_cols]


def load_and_clean_data(input_file: Path) -> pd.DataFrame:
    """Load all relevant sheets and produce one normalized cycle-level table."""
    print(f"正在加载数据：{input_file}")
    cached = load_cycles_cache(input_file)
    if cached is not None:
        print(f"  反应周期记录数：{len(cached)}")
        print(f"  日期范围：{cached['日期'].min().date()} ~ {cached['日期'].max().date()}")
        print(f"  炉号数量：{cached['炉号'].nunique()}")
        return cached

    frames: list[pd.DataFrame] = []
    skipped_tables: list[tuple[str, list[str]]] = []

    for sheet_index, sheet_name, df in iter_source_tables(input_file):
        if df.empty:
            continue

        missing = missing_required_columns(df)
        if missing:
            skipped_tables.append((sheet_name, missing))
            print(f"  跳过工作表 {sheet_name}，缺少列：{', '.join(missing)}")
            continue

        df = df.rename(
            columns={
                SOURCE_COLUMNS["date"]: "日期",
                SOURCE_COLUMNS["team"]: "班组",
                SOURCE_COLUMNS["furnace"]: "炉号",
                SOURCE_COLUMNS["reaction_time"]: "反应时间",
                SOURCE_COLUMNS["fault_time"]: "故障时间",
                SOURCE_COLUMNS["clean_empty_burn_time"]: "空烧时间",
                SOURCE_COLUMNS["output"]: "产量",
                SOURCE_COLUMNS["source_yield"]: "源表小时产能",
            }
        )
        df = df[
            ["日期", "班组", "炉号", "反应时间", "故障时间", "空烧时间", "产量", "源表小时产能"]
        ].copy()

        df["来源工作表"] = sheet_name
        df["来源行号"] = np.arange(2, len(df) + 2)
        df["工作表顺序"] = sheet_index
        sheet_line = production_line_from_sheet(sheet_name)

        df["班组"] = df["班组"].ffill()
        df["炉号"] = df["炉号"].astype(str).str.strip()
        df = df[(df["炉号"].notna()) & (df["炉号"] != "") & (df["炉号"] != "nan")]
        df = df[~df["炉号"].str.contains("总计", na=False)]
        df["生产线"] = df["炉号"].map(lambda furnace: production_line_from_furnace(furnace, sheet_line))

        frames.append(df)

    if not frames:
        if skipped_tables:
            raise ValueError(format_missing_column_error(input_file, skipped_tables))
        raise ValueError("未读取到可分析的数据，请检查输入表结构或空工作表")

    raw = pd.concat(frames, ignore_index=True)

    for col in ["反应时间", "故障时间", "空烧时间", "产量", "源表小时产能"]:
        raw[col] = raw[col].map(clean_number)

    raw["日期"] = raw["日期"].map(convert_excel_date)
    raw = raw[raw["日期"].notna()].copy()
    raw["年月"] = raw["日期"].dt.to_period("M").astype(str)
    raw["月份"] = raw["日期"].dt.month

    raw["故障时间"] = raw["故障时间"].fillna(0)
    raw["空烧时间"] = raw["空烧时间"].fillna(0)
    raw["降清时间"] = raw["空烧时间"]
    raw["产率"] = raw["产量"] / raw["反应时间"]
    raw.loc[~np.isfinite(raw["产率"]), "产率"] = np.nan

    raw = raw.sort_values(["日期", "工作表顺序", "来源行号"]).reset_index(drop=True)
    cycles = build_reaction_cycles(raw)

    print(f"  清洗后原始记录数：{len(raw)}")
    print(f"  反应周期记录数：{len(cycles)}")
    print(f"  日期范围：{cycles['日期'].min().date()} ~ {cycles['日期'].max().date()}")
    print(f"  炉号数量：{cycles['炉号'].nunique()}")
    save_cycles_cache(input_file, cycles)
    return cycles


def resolve_furnaces(cycles: pd.DataFrame, selectors: Iterable[str] | None) -> list[str] | None:
    """Resolve exact furnace names or prefix selectors."""
    if selectors is None:
        return None

    all_furnaces = sorted(cycles["炉号"].dropna().unique())
    selected: list[str] = []
    for raw_part in selectors:
        for part in [item.strip() for item in raw_part.split(",") if item.strip()]:
            if part in all_furnaces:
                selected.append(part)
                continue

            matched = [furnace for furnace in all_furnaces if furnace.startswith(part)]
            if matched:
                selected.extend(matched)
            else:
                print(f"  警告：未匹配到炉号或前缀 {part}")

    deduped = list(dict.fromkeys(selected))
    return deduped or []


def select_cycles(
    cycles: pd.DataFrame,
    furnace_list: list[str] | None,
    log_label: str | None = None,
) -> pd.DataFrame:
    """Return cycle records for all furnaces or a selected furnace list."""
    if furnace_list is None:
        if log_label:
            print(f"\n{log_label}：全区，共 {cycles['炉号'].nunique()} 个炉号")
        return cycles.copy()

    filtered = cycles[cycles["炉号"].isin(furnace_list)].copy()
    if log_label:
        print(f"\n{log_label}：自选炉子 {len(furnace_list)} 个：{', '.join(furnace_list)}")
    if filtered.empty:
        raise ValueError("选中的炉号没有数据")
    return filtered


def scope_name(furnace_list: list[str] | None) -> str:
    if furnace_list is None:
        return "全区"
    furnace_list = sorted(furnace_list)
    if len(furnace_list) <= 6:
        return "_".join(safe_file_part(f) for f in furnace_list)
    digest = hashlib.sha1(",".join(furnace_list).encode("utf-8")).hexdigest()[:6]
    # 按前缀分组：字母开头取首字母，数字开头取前3字符
    groups: dict[str, list[str]] = {}
    for f in furnace_list:
        key = f[0] if f[0].isalpha() else f[:3] if len(f) >= 3 else f
        groups.setdefault(key, []).append(f)
    parts: list[str] = []
    for key, group in groups.items():
        if len(group) == 1:
            parts.append(safe_file_part(group[0]))
        else:
            numbers = []
            for f in group:
                # 从右往左找末尾数字部分，如 11A17 -> 17, E01 -> 01
                last_nondigit = len(f) - 1
                while last_nondigit >= 0 and f[last_nondigit].isdigit():
                    last_nondigit -= 1
                suffix = f[last_nondigit + 1:]
                numbers.append(int(suffix) if suffix else 0)
            numbers.sort()
            # 取第一个炉子的前缀
            sample = group[0]
            last_nondigit = len(sample) - 1
            while last_nondigit >= 0 and sample[last_nondigit].isdigit():
                last_nondigit -= 1
            prefix = sample[:last_nondigit + 1]
            parts.append(f"{prefix}{numbers[0]}-{numbers[-1]}")
    joined = "&".join(parts[:4])
    label = f"自选{len(furnace_list)}炉_{joined}" if len(parts) <= 4 else f"自选{len(furnace_list)}炉等"
    return f"{label}_{digest}"


def monthly_furnace_average(data: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        data.groupby(["年月", "月份", "炉号", "生产线"], as_index=False)
        .agg(
            反应周期数=("周期序号", "count"),
            平均反应时间=("反应时间", "mean"),
            平均空烧时间=("空烧时间", "mean"),
            平均降清时间=("降清时间", "mean"),
            平均故障时间=("故障时间", "mean"),
            总产量=("产量", "sum"),
            总反应时间=("反应时间", "sum"),
        )
        .sort_values(["年月", "生产线", "炉号"])
    )

    grouped["平均产率"] = weighted_yield(grouped["总产量"], grouped["总反应时间"])
    grouped = grouped.drop(columns=["总产量", "总反应时间"])
    value_cols = [f"平均{metric}" for metric in CYCLE_METRICS]
    grouped[value_cols] = grouped[value_cols].round(2)
    return grouped


def region_monthly_average(data: pd.DataFrame) -> pd.DataFrame:
    result = (
        data.groupby(["年月", "月份"], as_index=False)
        .agg(
            反应周期数=("周期序号", "count"),
            平均反应时间=("反应时间", "mean"),
            平均空烧时间=("空烧时间", "mean"),
            平均降清时间=("降清时间", "mean"),
            平均故障时间=("故障时间", "mean"),
            总产量=("产量", "sum"),
            总反应时间=("反应时间", "sum"),
        )
        .sort_values("年月")
    )
    result["平均产率"] = weighted_yield(result["总产量"], result["总反应时间"])
    result = result.drop(columns=["总产量", "总反应时间"])
    value_cols = [f"平均{metric}" for metric in CYCLE_METRICS]
    result[value_cols] = result[value_cols].round(2)
    return result


def rank_monthly_top_bottom(monthly_avg: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Rank furnaces within each month and metric."""
    metric_rules = {f"平均{metric}": metric != "产率" for metric in CYCLE_METRICS}

    results: dict[str, pd.DataFrame] = {}
    summary_frames: list[pd.DataFrame] = []

    for metric, ascending_for_top in metric_rules.items():
        top_frames: list[pd.DataFrame] = []
        bottom_frames: list[pd.DataFrame] = []

        for month, sub in monthly_avg.dropna(subset=[metric]).groupby("年月", sort=True):
            ranked = sub.sort_values(metric, ascending=ascending_for_top).reset_index(drop=True)
            count = len(ranked)
            if count == 0:
                continue

            n20 = max(1, math.ceil(count * RANKING_TOP_PERCENT))
            top = ranked.head(n20).copy()
            bottom = ranked.tail(n20).sort_values(metric, ascending=not ascending_for_top).copy()
            top["类型"] = "前20%"
            bottom["类型"] = "后20%"
            top["排名"] = np.arange(1, len(top) + 1)
            bottom["排名"] = np.arange(1, len(bottom) + 1)
            top["指标"] = metric
            bottom["指标"] = metric
            top_frames.append(top)
            bottom_frames.append(bottom)

        top_result = pd.concat(top_frames, ignore_index=True) if top_frames else pd.DataFrame()
        bottom_result = pd.concat(bottom_frames, ignore_index=True) if bottom_frames else pd.DataFrame()

        keep_cols = ["年月", "月份", "指标", "类型", "排名", "炉号", "生产线", "反应周期数", metric]
        if not top_result.empty:
            top_result = top_result[keep_cols].rename(columns={metric: "数值"})
            summary_frames.append(top_result)
        if not bottom_result.empty:
            bottom_result = bottom_result[keep_cols].rename(columns={metric: "数值"})
            summary_frames.append(bottom_result)

        short_metric = metric.replace("平均", "")
        results[f"{short_metric}_前20%"] = top_result
        results[f"{short_metric}_后20%"] = bottom_result

    results["前后20汇总"] = (
        pd.concat(summary_frames, ignore_index=True)
        if summary_frames
        else pd.DataFrame(columns=["年月", "月份", "指标", "类型", "排名", "炉号", "生产线", "反应周期数", "数值"])
    )
    return results


def write_dataframe(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def autosize_workbook(path: Path) -> None:
    """Apply basic formatting and column widths to generated workbooks."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            width = 10
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                width = max(width, min(24, len(str(value)) + 2))
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    workbook.save(path)


def furnace_stats(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> Path:
    data = select_cycles(cycles, furnace_list, log_label="炉子级统计")
    prefix = scope_name(furnace_list)

    cycle_detail = data.sort_values(["日期", "生产线", "炉号", "周期序号"]).copy()
    cycle_detail[["反应时间", "空烧时间", "降清时间", "故障时间", "产量", "产率", "源表小时产能"]] = (
        cycle_detail[["反应时间", "空烧时间", "降清时间", "故障时间", "产量", "产率", "源表小时产能"]].round(2)
    )

    monthly_avg = monthly_furnace_average(data)
    region_avg = region_monthly_average(data)
    rank_results = rank_monthly_top_bottom(monthly_avg)

    output_path = OUTPUT_DIR / f"炉子级统计_{prefix}.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_dataframe(writer, cycle_detail, "反应周期明细")
        write_dataframe(writer, monthly_avg, "月度炉子平均")
        write_dataframe(writer, region_avg, f"{prefix}各月平均")
        write_dataframe(writer, rank_results["前后20汇总"], "前后20汇总")
        for name, result in rank_results.items():
            if name == "前后20汇总":
                continue
            write_dataframe(writer, result, name)

    autosize_workbook(output_path)
    print(f"  炉子级统计已输出：{output_path}")
    print(f"  周期明细：{len(cycle_detail)} 条；月度炉子平均：{len(monthly_avg)} 条")
    return output_path


def plot_furnace_stats_chart(cycles: pd.DataFrame, output_path: Path | None = None, furnace_list: list[str] | None = None):
    data = select_cycles(cycles, furnace_list)
    prefix = scope_name(furnace_list)
    region_avg = region_monthly_average(data)
    if region_avg.empty:
        raise ValueError("没有可生成炉子级统计图的数据")

    if output_path:
        print(f"正在生成炉子级统计图：{output_path}")
    x_plot = np.arange(len(region_avg))
    x_labels = region_avg["年月"].tolist()

    fig, axes = plt.subplots(2, 1, figsize=(15, 8.5), sharex=True, gridspec_kw={"height_ratios": [1.15, 1]})
    _style_figure(fig, axes)

    time_colors = {
        "平均反应时间": CHART_COLORS["reaction"],
        "平均空烧时间": CHART_COLORS["empty_burn"],
        "平均降清时间": CHART_COLORS["clean"],
        "平均故障时间": CHART_COLORS["fault"],
    }
    for metric, color in time_colors.items():
        axes[0].plot(x_plot, region_avg[metric], marker="o", linewidth=2.0, markersize=5, label=metric, color=color)

    axes[0].set_title(f"{prefix}炉子级月度平均统计", fontsize=15, fontweight="bold")
    axes[0].set_ylabel("平均时间 (h)")
    _style_legend(axes[0].legend(loc="upper left", ncol=2))

    yield_bars = axes[1].bar(x_plot, region_avg["平均产率"], color=CHART_COLORS["production"], alpha=0.86, label="加权平均产率")
    if len(yield_bars) <= 12:
        axes[1].bar_label(yield_bars, fmt="%.1f", fontsize=9, padding=4)
    min_rate = configured_min_yield_rate()
    if min_rate is not None:
        axes[1].axhline(min_rate, color=CHART_COLORS["threshold"], linestyle="--", linewidth=1.4, alpha=0.85, label=f"最低产率阈值 {min_rate:g}")
    axes[1].set_ylabel("平均产率 (kg/h)")
    _style_legend(axes[1].legend(loc="upper left"))

    ax_count = axes[1].twinx()
    _style_axis(ax_count, grid_axis=None)
    ax_count.spines["right"].set_visible(True)
    ax_count.plot(x_plot, region_avg["反应周期数"], color=CHART_COLORS["neutral"], linestyle="--", marker="s", linewidth=1.8, label="反应周期数")
    ax_count.set_ylabel("反应周期数")
    _style_legend(ax_count.legend(loc="upper right"))

    axes[1].set_xticks(x_plot)
    axes[1].set_xticklabels(x_labels)
    axes[1].set_xlabel("年月")
    fig.tight_layout()
    result = _save_or_buffer_figure(fig, output_path)
    if output_path is not None:
        print(f"  炉子级统计图已保存：{output_path}")
    return result


def run_furnace_stats(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[Path, Path]:
    workbook = furnace_stats(cycles, furnace_list)
    prefix = scope_name(furnace_list)
    chart = plot_furnace_stats_chart(cycles, OUTPUT_DIR / f"炉子级统计图_{prefix}.png", furnace_list)
    return workbook, chart


def summary_by_day(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    data = select_cycles(cycles, furnace_list)
    agg = {
        "产量": "sum",
        "反应时间": "sum",
        "故障时间": "sum",
        "空烧时间": "sum",
        "降清时间": "sum",
    }
    daily_all = data.groupby("日期", as_index=False).agg(agg).sort_values("日期")
    daily_line = data.groupby(["日期", "生产线"], as_index=False).agg(agg).sort_values(["生产线", "日期"])

    rename_cols = {
        "产量": "总产量",
        "反应时间": "总反应时间",
        "故障时间": "总故障时间",
        "空烧时间": "总空烧时间",
        "降清时间": "总降清时间",
    }
    daily_all = daily_all.rename(columns=rename_cols)
    daily_line = daily_line.rename(columns=rename_cols)
    daily_all = add_weighted_average_yield(daily_all)
    daily_line = add_weighted_average_yield(daily_line)
    daily_all[SUMMARY_METRICS] = daily_all[SUMMARY_METRICS].round(2)
    daily_line[SUMMARY_METRICS] = daily_line[SUMMARY_METRICS].round(2)
    return daily_all, daily_line


def summary_by_month(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    data = select_cycles(cycles, furnace_list)
    agg = {
        "产量": "sum",
        "反应时间": "sum",
        "故障时间": "sum",
        "空烧时间": "sum",
        "降清时间": "sum",
    }
    monthly_all = data.groupby("年月", as_index=False).agg(agg).sort_values("年月")
    monthly_line = data.groupby(["年月", "生产线"], as_index=False).agg(agg).sort_values(["生产线", "年月"])

    rename_cols = {
        "产量": "总产量",
        "反应时间": "总反应时间",
        "故障时间": "总故障时间",
        "空烧时间": "总空烧时间",
        "降清时间": "总降清时间",
    }
    monthly_all = monthly_all.rename(columns=rename_cols)
    monthly_line = monthly_line.rename(columns=rename_cols)
    monthly_all = add_weighted_average_yield(monthly_all)
    monthly_line = add_weighted_average_yield(monthly_line)
    monthly_all[SUMMARY_METRICS] = monthly_all[SUMMARY_METRICS].round(2)
    monthly_line[SUMMARY_METRICS] = monthly_line[SUMMARY_METRICS].round(2)
    return monthly_all, monthly_line


def write_daily_summary(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[Path, pd.DataFrame, pd.DataFrame]:
    print("\n正在生成每日汇总...")
    daily_all, daily_line = summary_by_day(cycles, furnace_list)
    prefix = scope_name(furnace_list)

    output_path = OUTPUT_DIR / ("每日汇总.xlsx" if furnace_list is None else f"每日汇总_{prefix}.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_dataframe(writer, daily_all, f"{prefix}每日汇总")
        write_dataframe(writer, daily_line, "生产线每日汇总")
        for line in sorted(daily_line["生产线"].unique()):
            write_dataframe(writer, daily_line[daily_line["生产线"] == line], f"{line}每日汇总")

    autosize_workbook(output_path)
    print(f"  每日汇总已输出：{output_path}")
    return output_path, daily_all, daily_line


def write_monthly_summary(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[Path, pd.DataFrame, pd.DataFrame]:
    print("\n正在生成每月汇总...")
    monthly_all, monthly_line = summary_by_month(cycles, furnace_list)
    prefix = scope_name(furnace_list)

    output_path = OUTPUT_DIR / ("每月汇总.xlsx" if furnace_list is None else f"每月汇总_{prefix}.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_dataframe(writer, monthly_all, f"{prefix}每月汇总")
        write_dataframe(writer, monthly_line, "生产线每月汇总")
        for line in sorted(monthly_line["生产线"].unique()):
            write_dataframe(writer, monthly_line[monthly_line["生产线"] == line], f"{line}每月汇总")

    autosize_workbook(output_path)
    print(f"  每月汇总已输出：{output_path}")
    return output_path, monthly_all, monthly_line


def plot_summary_trend(summary: pd.DataFrame, x_col: str, title: str, output_path: Path | None = None):
    if output_path:
        print(f"正在生成趋势图：{output_path}")
    fig, axes = plt.subplots(2, 1, figsize=(16, 9), sharex=True, gridspec_kw={"height_ratios": [1, 1.2]})
    _style_figure(fig, axes)

    x_values = summary[x_col]
    if x_col == "日期":
        x_plot = pd.to_datetime(x_values)
    else:
        x_plot = np.arange(len(summary))

    bars = axes[0].bar(x_plot, summary["总产量"], color=CHART_COLORS["production"], alpha=0.86, label="总产量")
    if len(bars) <= 12:
        axes[0].bar_label(bars, fmt="%.0f", fontsize=9, padding=4)
    axes[0].set_ylabel("总产量")
    axes[0].set_title(title, fontsize=15, fontweight="bold")
    _style_legend(axes[0].legend(loc="upper left"))

    if "平均产率" not in summary.columns and {"总产量", "总反应时间"} <= set(summary.columns):
        summary = add_weighted_average_yield(summary.copy())
    if "平均产率" in summary.columns:
        ax_yield = axes[0].twinx()
        _style_axis(ax_yield, grid_axis=None)
        ax_yield.spines["right"].set_visible(True)
        ax_yield.plot(x_plot, summary["平均产率"], color=CHART_COLORS["yield"], marker="o", linewidth=2.0, markersize=4, label="加权平均产率")
        min_rate = configured_min_yield_rate()
        if min_rate is not None:
            ax_yield.axhline(min_rate, color=CHART_COLORS["threshold"], linestyle="--", linewidth=1.3, alpha=0.75, label=f"最低产率阈值 {min_rate:g}")
        ax_yield.set_ylabel("平均产率 (kg/h)")
        _style_legend(ax_yield.legend(loc="upper right"))

    colors = {
        "总反应时间": CHART_COLORS["reaction"],
        "总故障时间": CHART_COLORS["fault"],
        "总空烧时间": CHART_COLORS["empty_burn"],
        "总降清时间": CHART_COLORS["clean"],
    }
    for metric in TIME_SUMMARY_METRICS:
        axes[1].plot(x_plot, summary[metric], marker="o", linewidth=1.8, markersize=4, label=metric, color=colors[metric])

    axes[1].set_ylabel("时间 (h)")
    _style_legend(axes[1].legend(loc="upper left", ncol=2))

    if x_col == "日期":
        _format_date_axis(axes[1], max_ticks=10)
        fig.autofmt_xdate(rotation=45)
    else:
        axes[1].set_xticks(x_plot)
        axes[1].set_xticklabels(summary[x_col].tolist())

    axes[1].set_xlabel(x_col)
    fig.tight_layout()
    result = _save_or_buffer_figure(fig, output_path)
    if output_path is not None:
        print(f"  趋势图已保存：{output_path}")
    return result


def safe_file_part(value: object) -> str:
    text = str(value).strip()
    for char in '<>:"/\\|?*':
        text = text.replace(char, "_")
    text = "_".join(text.split())
    return text or "未命名"


def furnace_daily_trend_data(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> pd.DataFrame:
    data = select_cycles(cycles, furnace_list)
    columns = [
        "日期",
        "年月",
        "生产线",
        "炉号",
        "班组",
        "反应时间",
        "空烧时间",
        "降清时间",
        "故障时间",
        "产量",
        "产率",
        "来源工作表",
        "来源行号",
    ]
    result = data[columns].sort_values(["生产线", "炉号", "日期"]).copy()
    numeric_cols = ["反应时间", "空烧时间", "降清时间", "故障时间", "产量", "产率"]
    result[numeric_cols] = result[numeric_cols].round(2)
    return result


def furnace_daily_trend_summary(trend_data: pd.DataFrame) -> pd.DataFrame:
    summary = (
        trend_data.groupby(["生产线", "炉号"], as_index=False)
        .agg(
            记录天数=("日期", "nunique"),
            总产量=("产量", "sum"),
            总反应时间=("反应时间", "sum"),
            总故障时间=("故障时间", "sum"),
            总空烧时间=("空烧时间", "sum"),
            总降清时间=("降清时间", "sum"),
        )
        .sort_values(["生产线", "炉号"])
    )
    summary = add_weighted_average_yield(summary)
    value_cols = ["总产量", "总反应时间", "总故障时间", "总空烧时间", "总降清时间", "平均产率"]
    summary[value_cols] = summary[value_cols].round(2)
    return summary


def write_furnace_daily_trend_data(
    cycles: pd.DataFrame, furnace_list: list[str] | None = None
) -> tuple[Path, pd.DataFrame]:
    print("\n正在生成单炉每日趋势数据...")
    trend_data = furnace_daily_trend_data(cycles, furnace_list)
    summary = furnace_daily_trend_summary(trend_data)
    prefix = scope_name(furnace_list)
    output_path = OUTPUT_DIR / f"单炉每日趋势数据_{prefix}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_dataframe(writer, trend_data, "单炉每日明细")
        write_dataframe(writer, summary, "单炉汇总")

    autosize_workbook(output_path)
    print(f"  单炉每日趋势数据已输出：{output_path}")
    return output_path, trend_data


def detect_anomalies(cycles: pd.DataFrame, sigma: float | None = None) -> pd.DataFrame:
    """Detect low-yield cycles by furnace using sigma and configured minimum yield thresholds."""
    threshold_sigma = float(sigma if sigma is not None else ALERT_THRESHOLDS.get("anomaly_sigma", 2.0))
    min_yield_rate = configured_min_yield_rate()
    results: list[pd.DataFrame] = []

    for furnace, sub in cycles.dropna(subset=["产率"]).groupby("炉号", sort=True):
        mean_value = sub["产率"].mean()
        std_value = sub["产率"].std()

        sigma_mask = pd.Series(False, index=sub.index)
        sigma_threshold = np.nan
        deviation = pd.Series(np.nan, index=sub.index)
        if not pd.isna(std_value) and std_value != 0:
            sigma_threshold = mean_value - threshold_sigma * std_value
            sigma_mask = sub["产率"] < sigma_threshold
            deviation = (mean_value - sub["产率"]) / std_value

        min_yield_mask = pd.Series(False, index=sub.index)
        if min_yield_rate is not None:
            min_yield_mask = sub["产率"] < min_yield_rate

        anomaly_mask = sigma_mask | min_yield_mask
        if not anomaly_mask.any():
            continue

        anomalies = sub[anomaly_mask].copy()

        def describe_anomaly(row_index: object) -> str:
            reasons: list[str] = []
            if bool(sigma_mask.loc[row_index]):
                reasons.append("低产率")
            if bool(min_yield_mask.loc[row_index]):
                reasons.append("低于最低产率阈值")
            return "；".join(reasons)

        def effective_threshold(row_index: object) -> float:
            thresholds: list[float] = []
            if bool(sigma_mask.loc[row_index]) and np.isfinite(sigma_threshold):
                thresholds.append(float(sigma_threshold))
            if bool(min_yield_mask.loc[row_index]) and min_yield_rate is not None:
                thresholds.append(float(min_yield_rate))
            return round(max(thresholds), 2) if thresholds else np.nan

        anomalies["异常类型"] = [describe_anomaly(index) for index in anomalies.index]
        anomalies["产率均值"] = round(mean_value, 2)
        anomalies["产率标准差"] = round(std_value, 2)
        anomalies["异常阈值"] = [effective_threshold(index) for index in anomalies.index]
        anomalies["偏差σ"] = deviation.loc[anomalies.index].round(2)
        results.append(anomalies)

    columns = [
        "日期",
        "年月",
        "生产线",
        "炉号",
        "班组",
        "异常类型",
        "产量",
        "反应时间",
        "故障时间",
        "空烧时间",
        "降清时间",
        "产率",
        "产率均值",
        "产率标准差",
        "异常阈值",
        "偏差σ",
        "来源工作表",
        "来源行号",
    ]
    if not results:
        return pd.DataFrame(columns=columns)

    output = pd.concat(results, ignore_index=True)
    output = output[columns].sort_values(["偏差σ", "日期"], ascending=[False, True])
    numeric_cols = ["产量", "反应时间", "故障时间", "空烧时间", "降清时间", "产率"]
    output[numeric_cols] = output[numeric_cols].round(2)
    return output


def write_anomaly_report(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> Path:
    print("\n正在生成异常检测报告...")
    data = select_cycles(cycles, furnace_list)
    anomalies = detect_anomalies(data)
    prefix = scope_name(furnace_list)
    output_path = OUTPUT_DIR / f"异常检测报告_{prefix}.xlsx"

    by_furnace = (
        anomalies.groupby(["生产线", "炉号"], as_index=False)
        .agg(异常次数=("日期", "count"), 最大偏差σ=("偏差σ", "max"), 最低产率=("产率", "min"))
        .sort_values(["异常次数", "最大偏差σ"], ascending=[False, False])
        if not anomalies.empty
        else pd.DataFrame(columns=["生产线", "炉号", "异常次数", "最大偏差σ", "最低产率"])
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_dataframe(writer, anomalies, "异常周期明细")
        write_dataframe(writer, by_furnace, "炉号异常汇总")

    autosize_workbook(output_path)
    print(f"  异常检测报告已输出：{output_path}，异常周期 {len(anomalies)} 条")
    return output_path


def yield_anomaly_mask(series: pd.Series) -> pd.Series:
    mask = pd.Series(False, index=series.index)
    min_yield_rate = configured_min_yield_rate()
    if min_yield_rate is not None:
        mask = mask | (series < min_yield_rate)

    clean = series.dropna()
    if len(clean) < 3:
        return mask.fillna(False)

    mean_value = clean.mean()
    std_value = clean.std()
    if pd.isna(std_value) or std_value == 0:
        return mask.fillna(False)

    sigma = float(ALERT_THRESHOLDS.get("anomaly_sigma", 2.0))
    return (mask | (series < (mean_value - sigma * std_value))).fillna(False)


def detect_cycle_boundaries(data: pd.DataFrame, threshold: float = 20.0) -> pd.DataFrame:
    """Detect cycle boundaries where cumulative 空烧+故障 >= threshold hours."""
    df = data.sort_values("日期").copy()
    df["空烧故障合计"] = df["空烧时间"].fillna(0) + df["故障时间"].fillna(0)
    cumulative = 0.0
    cycle_id = 1
    boundaries = []
    for idx, row in df.iterrows():
        cumulative += row["空烧故障合计"]
        df.at[idx, "累计空闲"] = cumulative
        df.at[idx, "周期编号"] = cycle_id
        boundaries.append(False)
        if cumulative >= threshold:
            boundaries[-1] = True
            cumulative = 0.0
            cycle_id += 1
    df["周期边界"] = boundaries

    cycle_stats = (
        df.groupby("周期编号")
        .agg(
            周期开始=("日期", "min"),
            周期结束=("日期", "max"),
            周期天数=("日期", lambda x: (x.max() - x.min()).days + 1),
            周期反应时间=("反应时间", "sum"),
            周期产量=("产量", "sum"),
            周期空烧时间=("空烧时间", "sum"),
            周期故障时间=("故障时间", "sum"),
            记录数=("周期编号", "count"),
        )
        .assign(周期产率=lambda x: (x["周期产量"] / x["周期反应时间"]).round(2))
    )
    return cycle_stats


def plot_single_furnace_daily_trend(furnace_data: pd.DataFrame, output_path: Path | None = None):
    furnace = furnace_data["炉号"].iloc[0]
    line = furnace_data["生产线"].iloc[0]
    data = furnace_data.sort_values("日期").copy()

    full_dates = pd.date_range(data["日期"].min(), data["日期"].max(), freq="D")
    plot_data = data.set_index("日期").reindex(full_dates)
    for col in ["产量", "反应时间", "空烧时间", "降清时间", "故障时间"]:
        plot_data[col] = plot_data[col].fillna(0)

    fig, axes = plt.subplots(2, 1, figsize=(14, 8), sharex=True, gridspec_kw={"height_ratios": [1, 1.15]})
    _style_figure(fig, axes)
    x_values = plot_data.index

    bars = axes[0].bar(x_values, plot_data["产量"], color=CHART_COLORS["production"], alpha=0.86, label="产量")
    if len(bars) <= 12:
        axes[0].bar_label(bars, fmt="%.0f", fontsize=8, padding=3)
    axes[0].set_ylabel("产量")
    _style_legend(axes[0].legend(loc="upper left"))

    ax_yield = axes[0].twinx()
    _style_axis(ax_yield, grid_axis=None)
    ax_yield.spines["right"].set_visible(True)
    yield_series = plot_data["产率"].astype(float)
    ax_yield.plot(x_values, yield_series, color=CHART_COLORS["yield"], linewidth=1.7, marker="o", markersize=3, label="产率")

    moving_avg = yield_series.rolling(window=7, min_periods=3).mean()
    ax_yield.plot(x_values, moving_avg, color=CHART_COLORS["highlight_alt"], linewidth=2.0, label="产率7日均线")

    valid = yield_series.dropna()
    if len(valid) >= 2:
        x_index = np.arange(len(yield_series))
        valid_mask = yield_series.notna().to_numpy()
        slope, intercept = np.polyfit(x_index[valid_mask], yield_series[valid_mask], 1)
        trend = slope * x_index + intercept
        ax_yield.plot(x_values, trend, color=CHART_COLORS["neutral"], linestyle="--", linewidth=1.6, label=f"趋势线({slope:.2f}/天)")

    anomaly_mask = yield_anomaly_mask(yield_series)
    if anomaly_mask.any():
        ax_yield.scatter(
            x_values[anomaly_mask.to_numpy()],
            yield_series[anomaly_mask],
            color=CHART_COLORS["threshold"],
            s=36,
            zorder=5,
            label="异常低产",
        )

    ax_yield.set_ylabel("产率 (kg/h)")
    _style_legend(ax_yield.legend(loc="upper right"))

    colors = {
        "反应时间": CHART_COLORS["reaction"],
        "故障时间": CHART_COLORS["fault"],
        "空烧时间": CHART_COLORS["empty_burn"],
        "降清时间": CHART_COLORS["clean"],
    }
    for metric, color in colors.items():
        axes[1].plot(x_values, plot_data[metric], marker="o", markersize=3, linewidth=1.6, label=metric, color=color)

    fault_threshold = float(ALERT_THRESHOLDS.get("max_fault_hours_per_day", 24))
    heavy_fault_days = plot_data.index[plot_data["故障时间"] > fault_threshold]
    for day in heavy_fault_days:
        axes[0].axvline(day, color=CHART_COLORS["threshold"], alpha=0.16, linewidth=1.2)
        axes[1].axvline(day, color=CHART_COLORS["threshold"], alpha=0.18, linewidth=1.2)

    # 周期边界标注（空烧+故障 >= 20h）
    try:
        cycle_stats = detect_cycle_boundaries(data)
        for _, row in cycle_stats.iterrows():
            end_date = row["周期结束"]
            if end_date in plot_data.index:
                axes[0].axvline(end_date, color=CHART_COLORS["highlight"], alpha=0.35, linewidth=1.8, linestyle="--")
                axes[1].axvline(end_date, color=CHART_COLORS["highlight"], alpha=0.35, linewidth=1.8, linestyle="--")
        # 左上角周期统计
        total_cycles = len(cycle_stats)
        total_time = cycle_stats["周期反应时间"].sum()
        avg_yield = (cycle_stats["周期产量"].sum() / total_time).round(1) if total_time > 0 else 0
        axes[0].text(0.02, 0.95, f"周期数: {total_cycles}  总反应时间: {total_time:.0f}h  总平均产率: {avg_yield}",
                     transform=axes[0].transAxes, fontsize=8, verticalalignment="top",
                     bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.86, edgecolor=CHART_BORDER))
    except Exception:
        pass

    axes[1].set_ylabel("时间 (h)")
    _style_legend(axes[1].legend(loc="upper left", ncol=2))
    _format_date_axis(axes[1], max_ticks=10)
    axes[1].set_xlabel("日期")

    title = f"{line} {furnace} 单炉每日趋势"
    fig.suptitle(title, fontsize=15, fontweight="bold")
    fig.autofmt_xdate(rotation=45)
    fig.tight_layout(rect=(0, 0, 1, 0.96))
    return _save_or_buffer_figure(fig, output_path)


def format_duration(seconds: float) -> str:
    if seconds < 60:
        return f"{math.ceil(seconds)} 秒"
    return f"{math.ceil(seconds / 60)} 分钟"


def plot_furnace_daily_trends(trend_data: pd.DataFrame, prefix: str) -> tuple[Path, list[Path]]:
    chart_dir = OUTPUT_DIR / f"单炉每日趋势图_{prefix}"
    chart_dir.mkdir(parents=True, exist_ok=True)
    for old_chart in chart_dir.glob("*.png"):
        old_chart.unlink()
    chart_paths: list[Path] = []

    grouped = list(trend_data.groupby(["生产线", "炉号"], sort=True))
    total = len(grouped)
    print(f"正在生成单炉每日趋势图，共 {total} 个炉号...")
    if total >= 20:
        estimated = format_duration(total * 0.8)
        print(f"  提示：需要逐炉生成 PNG，预计耗时约 {estimated}，请等待进度提示。")

    for index, ((line, furnace), sub) in enumerate(grouped, start=1):
        file_name = f"{safe_file_part(line)}_{safe_file_part(furnace)}_每日趋势.png"
        path = plot_single_furnace_daily_trend(sub, chart_dir / file_name)
        chart_paths.append(path)
        if index == total or index % 20 == 0:
            print(f"  已生成 {index}/{total} 张")

    print(f"  单炉每日趋势图目录：{chart_dir}")
    return chart_dir, chart_paths


def run_furnace_daily_trends(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[Path, Path, list[Path]]:
    prefix = scope_name(furnace_list)
    workbook, trend_data = write_furnace_daily_trend_data(cycles, furnace_list)
    chart_dir, chart_paths = plot_furnace_daily_trends(trend_data, prefix)
    return workbook, chart_dir, chart_paths


def plot_furnace_yield_comparison(cycles: pd.DataFrame, output_path: Path | None = None,
                                   furnace_list: list[str] | None = None,
                                   start_date=None, end_date=None) -> Path | io.BytesIO:
    """Bar chart: x=furnace, y=average yield rate over selected date range."""
    data = select_cycles(cycles, furnace_list)
    if start_date is not None:
        data = data[data["日期"] >= pd.Timestamp(start_date)]
    if end_date is not None:
        data = data[data["日期"] <= pd.Timestamp(end_date)]
    avg = data.groupby("炉号")["产率"].mean().sort_values().dropna()
    if avg.empty:
        raise ValueError("所选范围无有效产率数据")

    fig_width = min(18, max(10, len(avg) * 0.12 + 8))
    fig_height = min(24, max(6, len(avg) * 0.22 + 2.8))
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    _style_axis(ax, grid_axis="x")
    colors = [CHART_COLORS["low"] if v < avg.median() else CHART_COLORS["ok"] for v in avg.values]
    ax.barh(avg.index, avg.values, color=colors, alpha=0.88)
    ax.axvline(avg.mean(), color=CHART_COLORS["reaction"], linestyle="--", linewidth=1.5, label=f"总平均 {avg.mean():.1f}")
    ax.set_xlabel("平均产率 (kg/h)")
    ax.set_title(f"炉号平均产率对比（{len(avg)} 个炉子）", fontsize=14, fontweight="bold")
    _style_legend(ax.legend())
    fig.tight_layout()

    return _save_or_buffer_figure(fig, output_path)


def plot_daily_furnace_yield_comparison(cycles: pd.DataFrame, output_path: Path | None = None,
                                         furnace_list: list[str] | None = None) -> Path | io.BytesIO:
    """Line chart: each furnace's daily yield rate over time for comparison."""
    data = select_cycles(cycles, furnace_list)
    daily = data.pivot_table(index="日期", columns="炉号", values="产率", aggfunc="mean")
    if daily.empty:
        raise ValueError("无有效产率数据")

    fig, ax = plt.subplots(figsize=(16, 7))
    _style_axis(ax, grid_axis="y")
    colors = plt.cm.tab20(np.linspace(0, 1, min(20, len(daily.columns))))
    plotted = 0
    for i, furnace in enumerate(daily.columns):
        series = daily[furnace].dropna()
        if len(series) >= 1:
            ax.plot(series.index, series.values, color=colors[i % 20], linewidth=1.2,
                    marker=".", markersize=2, alpha=0.8, label=furnace)
            plotted += 1

    ax.set_title(f"每日多炉产率对比（{len(daily.columns)} 个炉子）", fontsize=14, fontweight="bold")
    ax.set_ylabel("产率 (kg/h)")
    legend_cols = min(8, max(1, math.ceil(max(plotted, 1) / 12)))
    _style_legend(ax.legend(loc="upper left", ncol=legend_cols, fontsize=7, markerscale=2))
    _format_date_axis(ax, max_ticks=10)
    fig.autofmt_xdate(rotation=45)
    fig.tight_layout()

    return _save_or_buffer_figure(fig, output_path)


def plot_yield_heatmap(cycles: pd.DataFrame, output_path: Path | None = None,
                        furnace_list: list[str] | None = None,
                        value_col: str = "产率") -> Path | io.BytesIO:
    """产率热力图：X=日期, Y=炉号, 颜色=产率，替代拥挤折线"""
    data = select_cycles(cycles, furnace_list)
    pivot = data.pivot_table(index="炉号", columns="日期", values=value_col, aggfunc="mean")
    pivot = pivot.sort_index()

    fig_h = max(7, min(28, 0.22 * len(pivot.index) + 3))
    fig_w = max(12, min(28, 0.18 * len(pivot.columns) + 5))
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    heatmap = ax.imshow(pivot.to_numpy(), aspect="auto", cmap="RdYlGn", interpolation="bilinear")

    y_step = max(1, len(pivot.index) // 40)
    ax.set_yticks(np.arange(0, len(pivot.index), y_step))
    ax.set_yticklabels(pivot.index[::y_step], fontsize=8)
    x_step = max(1, len(pivot.columns) // 14)
    ax.set_xticks(np.arange(0, len(pivot.columns), x_step))
    ax.set_xticklabels([pd.Timestamp(pivot.columns[i]).strftime("%m/%d") for i in range(0, len(pivot.columns), x_step)],
                       rotation=45, ha="right", fontsize=8)

    cbar = fig.colorbar(heatmap, ax=ax)
    cbar.set_label(value_col)
    ax.set_title(f"多炉{value_col}热力图（{len(pivot.index)}炉 × {len(pivot.columns)}天）", fontsize=14, fontweight="bold")
    ax.set_xlabel("日期")
    ax.set_ylabel("炉号")
    _style_axis(ax, grid_axis=None)
    fig.tight_layout()
    return _save_or_buffer_figure(fig, output_path)


def plot_cycle_heatmap(cycles: pd.DataFrame, output_path: Path | None = None,
                        furnace_list: list[str] | None = None,
                        value_col: str = "反应时间") -> Path | io.BytesIO:
    """周期热力图：X=日期, Y=炉号, 颜色=周期指标，直观发现异常模式"""
    data = select_cycles(cycles, furnace_list)
    pivot = data.pivot_table(index="炉号", columns="日期", values=value_col, aggfunc="sum")
    pivot = pivot.sort_index()

    fig_h = max(7, min(28, 0.22 * len(pivot.index) + 3))
    fig_w = max(12, min(28, 0.18 * len(pivot.columns) + 5))
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    heatmap = ax.imshow(pivot.to_numpy(), aspect="auto", cmap="YlOrRd", interpolation="bilinear")

    y_step = max(1, len(pivot.index) // 40)
    ax.set_yticks(np.arange(0, len(pivot.index), y_step))
    ax.set_yticklabels(pivot.index[::y_step], fontsize=8)
    x_step = max(1, len(pivot.columns) // 14)
    ax.set_xticks(np.arange(0, len(pivot.columns), x_step))
    ax.set_xticklabels([pd.Timestamp(pivot.columns[i]).strftime("%m/%d") for i in range(0, len(pivot.columns), x_step)],
                       rotation=45, ha="right", fontsize=8)

    cbar = fig.colorbar(heatmap, ax=ax)
    cbar.set_label(f"{value_col} (h)" if "时间" in value_col else value_col)
    ax.set_title(f"周期{value_col}热力图（{len(pivot.index)}炉 × {len(pivot.columns)}天）", fontsize=14, fontweight="bold")
    ax.set_xlabel("日期")
    ax.set_ylabel("炉号")
    _style_axis(ax, grid_axis=None)
    fig.tight_layout()
    return _save_or_buffer_figure(fig, output_path)


def plot_cycle_time_distribution(cycles: pd.DataFrame, output_path: Path | None = None,
                                   furnace_list: list[str] | None = None) -> Path | io.BytesIO:
    """周期工作时间分布：直方图 + 箱线图"""
    data = select_cycles(cycles, furnace_list)
    stats = detect_cycle_boundaries(data)
    if stats.empty:
        raise ValueError("无有效周期数据")

    fig, axes = plt.subplots(2, 2, figsize=(14, 9))
    _style_figure(fig, axes)
    times = stats["周期反应时间"].dropna()

    # 1. 反应时间直方图
    axes[0, 0].hist(times, bins=min(30, max(5, len(times) // 5)), color=CHART_COLORS["production"], alpha=0.86, edgecolor="white")
    axes[0, 0].axvline(times.mean(), color=CHART_COLORS["threshold"], linestyle="--", linewidth=1.5, label=f"均值 {times.mean():.1f}h")
    axes[0, 0].axvline(times.median(), color=CHART_COLORS["reaction"], linestyle="--", linewidth=1.5, label=f"中位 {times.median():.1f}h")
    axes[0, 0].set_title("周期反应时间分布", fontsize=13, fontweight="bold")
    axes[0, 0].set_xlabel("反应时间 (h)")
    axes[0, 0].set_ylabel("周期数")
    _style_legend(axes[0, 0].legend(fontsize=8))

    # 2. 周期天数直方图
    days = stats["周期天数"].dropna()
    axes[0, 1].hist(days, bins=min(20, max(3, len(days) // 5)), color=CHART_COLORS["empty_burn"], alpha=0.86, edgecolor="white")
    axes[0, 1].axvline(days.mean(), color=CHART_COLORS["threshold"], linestyle="--", linewidth=1.5, label=f"均值 {days.mean():.1f}天")
    axes[0, 1].set_title("周期天数分布", fontsize=13, fontweight="bold")
    axes[0, 1].set_xlabel("天数")
    axes[0, 1].set_ylabel("周期数")
    _style_legend(axes[0, 1].legend(fontsize=8))

    # 3. 产率分布直方图
    yields = stats["周期产率"].dropna()
    axes[1, 0].hist(yields, bins=min(30, max(5, len(yields) // 5)), color=CHART_COLORS["clean"], alpha=0.86, edgecolor="white")
    axes[1, 0].axvline(yields.mean(), color=CHART_COLORS["threshold"], linestyle="--", linewidth=1.5, label=f"均值 {yields.mean():.1f}")
    axes[1, 0].set_title("周期产率分布", fontsize=13, fontweight="bold")
    axes[1, 0].set_xlabel("产率 (kg/h)")
    axes[1, 0].set_ylabel("周期数")
    _style_legend(axes[1, 0].legend(fontsize=8))

    # 4. 统计摘要文本
    axes[1, 1].axis("off")
    total_cycles = len(stats)
    summary_lines = [
        f"周期总数: {total_cycles}",
        f"平均反应时间: {times.mean():.1f} h",
        f"反应时间中位: {times.median():.1f} h",
        f"反应时间标准差: {times.std():.1f} h",
        f"最短/最长反应时间: {times.min():.1f} / {times.max():.1f} h",
        "",
        f"平均周期天数: {days.mean():.1f} 天",
        f"最短/最长天数: {days.min():.0f} / {days.max():.0f} 天",
        "",
        f"平均周期产率: {yields.mean():.1f} kg/h",
        f"周期总产量: {stats['周期产量'].sum():.0f} kg",
    ]
    for i, line in enumerate(summary_lines):
        axes[1, 1].text(0.05, 0.95 - i * 0.08, line, transform=axes[1, 1].transAxes,
                        fontsize=11, color=CHART_TEXT,
                        fontweight="bold" if i == 0 else "normal")

    fig.suptitle(f"周期工作时间分布（{total_cycles} 个周期）", fontsize=15, fontweight="bold", y=0.99)
    fig.tight_layout(rect=(0, 0, 1, 0.95))

    return _save_or_buffer_figure(fig, output_path)


def plot_3d_yield_comparison(cycles: pd.DataFrame, output_path: Path | None = None,
                              furnace_list: list[str] | None = None,
                              start_date=None, end_date=None,
                              highlight_furnace: str | None = None,
                              highlight_date=None) -> Path | io.BytesIO:
    """3D bar chart: X=date, Y=furnace, Z=yield rate. Supports highlight."""
    from mpl_toolkits.mplot3d import Axes3D  # noqa: F401
    data = select_cycles(cycles, furnace_list)
    if start_date is not None:
        data = data[data["日期"] >= pd.Timestamp(start_date)]
    if end_date is not None:
        data = data[data["日期"] <= pd.Timestamp(end_date)]

    dates = [pd.Timestamp(value).normalize() for value in sorted(data["日期"].dropna().unique())]
    furnaces = sorted(data["炉号"].dropna().unique())
    if len(dates) < 2 or len(furnaces) < 2:
        raise ValueError("需要至少2个日期和2个炉子来生成三维图")

    date_map = {d: i for i, d in enumerate(dates)}
    daily = data.pivot_table(index="日期", columns="炉号", values="产率", aggfunc="mean").reindex(dates)[furnaces]

    fig = plt.figure(figsize=(16, 9))
    fig.patch.set_facecolor(CHART_SURFACE)
    ax = fig.add_subplot(111, projection="3d")
    ax.view_init(elev=28, azim=-55)

    xpos, ypos = np.meshgrid(np.arange(len(dates)), np.arange(len(furnaces)), indexing="ij")
    xpos, ypos = xpos.ravel(), ypos.ravel()
    zpos = np.zeros_like(xpos)
    dz = np.array([daily.iloc[xi, yi] if pd.notna(daily.iloc[xi, yi]) else 0
                   for xi, yi in zip(xpos, ypos)])
    dx = dy = 0.65

    # 梯度颜色映射
    valid_dz = dz[dz > 0]
    vmin, vmax = (np.percentile(valid_dz, 5), np.percentile(valid_dz, 95)) if len(valid_dz) > 0 else (0, 100)
    cmap = plt.cm.viridis
    hl_date_idx = -1
    if highlight_date:
        hl_date_idx = date_map.get(pd.Timestamp(highlight_date).normalize(), -1)

    bar_colors = []
    for i, (xi, yi, z) in enumerate(zip(xpos, ypos, dz)):
        is_hl_furn = highlight_furnace and furnaces[yi] == highlight_furnace
        is_hl_date = xi == hl_date_idx
        if z <= 0:
            bar_colors.append("#EAEAEA")
        elif is_hl_furn and is_hl_date:
            bar_colors.append(CHART_COLORS["threshold"])
        elif is_hl_furn:
            bar_colors.append(CHART_COLORS["highlight"])
        elif is_hl_date:
            bar_colors.append(CHART_COLORS["highlight_alt"])
        else:
            bar_colors.append(cmap((z - vmin) / (vmax - vmin + 0.001)))

    ax.bar3d(xpos, ypos, zpos, dx, dy, dz, color=bar_colors, alpha=0.92, shade=True, edgecolor="white", linewidth=0.15)

    # 每7天一个刻度，避免拥挤
    tick_step = max(1, len(dates) // 12)
    ax.set_xticks(np.arange(0, len(dates), tick_step))
    ax.set_xticklabels([dates[i].strftime("%m/%d") for i in range(0, len(dates), tick_step)], fontsize=8, rotation=30)
    # 每2个炉号一个刻度
    y_step = max(1, len(furnaces) // 20)
    ax.set_yticks(np.arange(0, len(furnaces), y_step))
    ax.set_yticklabels([furnaces[i] for i in range(0, len(furnaces), y_step)], fontsize=8)
    ax.set_zlabel("产率 (kg/h)", fontsize=10, labelpad=12)
    ax.set_title(f"多炉每日产率对比（{len(dates)}天 × {len(furnaces)}炉）", fontsize=15, fontweight="bold", pad=20)
    ax.tick_params(colors=CHART_MUTED, labelsize=8)
    ax.xaxis.label.set_color(CHART_MUTED)
    ax.yaxis.label.set_color(CHART_MUTED)
    ax.zaxis.label.set_color(CHART_MUTED)
    ax.title.set_color(CHART_TEXT)

    ax.xaxis.pane.fill = False
    ax.yaxis.pane.fill = False
    ax.zaxis.pane.fill = False
    ax.xaxis.pane.set_edgecolor(CHART_BORDER)
    ax.yaxis.pane.set_edgecolor(CHART_BORDER)
    ax.zaxis.pane.set_edgecolor(CHART_BORDER)
    ax.grid(True, alpha=0.15)

    if highlight_furnace:
        ax.text2D(0.02, 0.97, f"◇ 高亮炉子: {highlight_furnace}", transform=ax.transAxes,
                  fontsize=10, color=CHART_COLORS["highlight"], fontweight="bold",
                  bbox=dict(boxstyle="round,pad=0.4", facecolor="white", alpha=0.88, edgecolor=CHART_BORDER))
    if highlight_date:
        ax.text2D(0.02, 0.92, f"◇ 高亮日期: {highlight_date}", transform=ax.transAxes,
                  fontsize=10, color=CHART_COLORS["highlight_alt"], fontweight="bold",
                  bbox=dict(boxstyle="round,pad=0.4", facecolor="white", alpha=0.88, edgecolor=CHART_BORDER))

    fig.subplots_adjust(left=0.03, right=0.98, bottom=0.07, top=0.92)
    return _save_or_buffer_figure(fig, output_path)


def fault_ranking(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> pd.DataFrame:
    data = select_cycles(cycles, furnace_list)
    flagged = data.assign(是否故障=data["故障时间"].fillna(0) > 0)
    ranking = (
        flagged.groupby(["生产线", "炉号"], as_index=False)
        .agg(
            故障总时间=("故障时间", "sum"),
            故障次数=("是否故障", "sum"),
            故障天数=("日期", lambda values: values[flagged.loc[values.index, "是否故障"]].nunique()),
            平均故障时间=("故障时间", "mean"),
            最大单日故障时间=("故障时间", "max"),
        )
        .sort_values(["故障总时间", "故障次数"], ascending=[False, False])
    )
    value_cols = ["故障总时间", "平均故障时间", "最大单日故障时间"]
    ranking[value_cols] = ranking[value_cols].round(2)
    return ranking


def fault_weekday_distribution(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> pd.DataFrame:
    data = select_cycles(cycles, furnace_list)
    fault_data = data[data["故障时间"].fillna(0) > 0].copy()
    if fault_data.empty:
        return pd.DataFrame(columns=["星期", "故障次数", "故障总时间"])

    weekday_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    fault_data["星期"] = fault_data["日期"].dt.dayofweek.map(lambda idx: weekday_names[idx])
    result = (
        fault_data.groupby("星期", as_index=False)
        .agg(故障次数=("故障时间", "count"), 故障总时间=("故障时间", "sum"))
    )
    result["星期序号"] = result["星期"].map({name: index for index, name in enumerate(weekday_names)})
    result = result.sort_values("星期序号").drop(columns="星期序号")
    result["故障总时间"] = result["故障总时间"].round(2)
    return result


def plot_fault_heatmap(cycles: pd.DataFrame, output_path: Path | None = None, furnace_list: list[str] | None = None):
    data = select_cycles(cycles, furnace_list)
    pivot = data.pivot_table(index="炉号", columns="日期", values="故障时间", aggfunc="sum", fill_value=0)
    pivot = pivot.sort_index()
    if pivot.empty:
        raise ValueError("无故障热力图数据")

    fig_height = max(7, min(28, 0.18 * len(pivot.index) + 3))
    fig_width = max(12, min(28, 0.16 * len(pivot.columns) + 5))
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    fig.patch.set_facecolor(CHART_SURFACE)
    heatmap = ax.imshow(pivot.to_numpy(), aspect="auto", cmap="YlOrRd")

    ax.set_title("故障时间热力图", fontsize=15, fontweight="bold", color=CHART_TEXT)
    ax.set_ylabel("炉号")
    ax.set_xlabel("日期")
    ax.tick_params(axis="both", colors=CHART_MUTED, labelsize=9)

    y_step = max(1, math.ceil(len(pivot.index) / 45))
    ax.set_yticks(np.arange(0, len(pivot.index), y_step))
    ax.set_yticklabels(pivot.index[::y_step])

    x_step = max(1, math.ceil(len(pivot.columns) / 14))
    x_positions = np.arange(0, len(pivot.columns), x_step)
    ax.set_xticks(x_positions)
    ax.set_xticklabels([pd.Timestamp(pivot.columns[i]).strftime("%m/%d") for i in x_positions], rotation=45, ha="right")

    for spine in ax.spines.values():
        spine.set_color(CHART_BORDER)
        spine.set_linewidth(0.8)
    cbar = fig.colorbar(heatmap, ax=ax)
    cbar.set_label("故障时间 (h)")
    cbar.ax.tick_params(colors=CHART_MUTED)
    fig.tight_layout()
    result = _save_or_buffer_figure(fig, output_path)
    if output_path is not None:
        print(f"  故障热力图已保存：{output_path}")
    return result


def write_fault_analysis_report(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> Path:
    print("\n正在生成故障分析报表...")
    prefix = scope_name(furnace_list)
    ranking = fault_ranking(cycles, furnace_list)
    weekday = fault_weekday_distribution(cycles, furnace_list)
    output_path = OUTPUT_DIR / f"故障分析_{prefix}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_dataframe(writer, ranking, "故障炉号排名")
        write_dataframe(writer, weekday, "故障星期分布")

    autosize_workbook(output_path)
    print(f"  故障分析已输出：{output_path}")
    return output_path


def run_fault_analysis(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[Path, Path]:
    print("\n正在生成故障分析...")
    prefix = scope_name(furnace_list)
    output_path = write_fault_analysis_report(cycles, furnace_list)
    heatmap_path = plot_fault_heatmap(cycles, OUTPUT_DIR / f"故障热力图_{prefix}.png", furnace_list)
    return output_path, heatmap_path


def detect_fault_warnings(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    data = select_cycles(cycles, furnace_list).copy()
    warning_hours = float(ALERT_THRESHOLDS.get("fault_warning_hours_per_day", 8))
    critical_hours = float(ALERT_THRESHOLDS.get("fault_critical_hours_per_day", 12))
    consecutive_days = int(ALERT_THRESHOLDS.get("consecutive_fault_days", 2))
    monthly_hours = float(ALERT_THRESHOLDS.get("monthly_fault_hours_warning", 24))

    warnings: list[dict] = []
    high_fault = data[data["故障时间"].fillna(0) >= warning_hours].copy()
    for _, row in high_fault.iterrows():
        fault_time = float(row["故障时间"])
        level = "严重" if fault_time >= critical_hours else "预警"
        warnings.append(
            {
                "预警级别": level,
                "预警类型": "单日故障超阈值",
                "日期": row["日期"],
                "年月": row["年月"],
                "生产线": row["生产线"],
                "炉号": row["炉号"],
                "故障时间": round(fault_time, 2),
                "阈值数值": critical_hours if level == "严重" else warning_hours,
                "阈值单位": "小时",
                "说明": f"单日故障时间 {fault_time:.2f} 小时",
                "来源工作表": row.get("来源工作表", ""),
                "来源行号": row.get("来源行号", ""),
            }
        )

    for (line, furnace), sub in data.sort_values("日期").groupby(["生产线", "炉号"], sort=True):
        fault_sub = sub[sub["故障时间"].fillna(0) > 0].copy()
        if fault_sub.empty:
            continue
        fault_sub["连续组"] = (fault_sub["日期"].diff().dt.days.ne(1)).cumsum()
        for _, group in fault_sub.groupby("连续组"):
            if len(group) >= consecutive_days:
                warnings.append(
                    {
                        "预警级别": "预警",
                        "预警类型": "连续故障",
                        "日期": group["日期"].iloc[-1],
                        "年月": group["年月"].iloc[-1],
                        "生产线": line,
                        "炉号": furnace,
                        "故障时间": round(group["故障时间"].sum(), 2),
                        "阈值数值": consecutive_days,
                        "阈值单位": "天",
                        "说明": f"连续 {len(group)} 天出现故障，区间 {group['日期'].iloc[0].date()} ~ {group['日期'].iloc[-1].date()}",
                        "来源工作表": join_unique(group["来源工作表"]),
                        "来源行号": join_unique(group["来源行号"]),
                    }
                )

    monthly = (
        data.groupby(["年月", "生产线", "炉号"], as_index=False)
        .agg(月故障总时间=("故障时间", "sum"), 故障天数=("故障时间", lambda values: (values.fillna(0) > 0).sum()))
    )
    monthly = monthly[monthly["月故障总时间"] >= monthly_hours]
    for _, row in monthly.iterrows():
        warnings.append(
            {
                "预警级别": "预警",
                "预警类型": "月累计故障超阈值",
                "日期": pd.NaT,
                "年月": row["年月"],
                "生产线": row["生产线"],
                "炉号": row["炉号"],
                "故障时间": round(float(row["月故障总时间"]), 2),
                "阈值数值": monthly_hours,
                "阈值单位": "小时",
                "说明": f"月累计故障 {row['月故障总时间']:.2f} 小时，故障天数 {int(row['故障天数'])} 天",
                "来源工作表": "",
                "来源行号": "",
            }
        )

    columns = ["预警级别", "预警类型", "日期", "年月", "生产线", "炉号", "故障时间", "阈值数值", "阈值单位", "说明", "来源工作表", "来源行号"]
    warning_df = pd.DataFrame(warnings, columns=columns)
    if not warning_df.empty:
        level_order = {"严重": 0, "预警": 1}
        warning_df["级别排序"] = warning_df["预警级别"].map(level_order).fillna(9)
        warning_df = warning_df.sort_values(["级别排序", "年月", "炉号", "预警类型"]).drop(columns="级别排序")

    summary = (
        warning_df.groupby(["生产线", "炉号", "预警级别"], as_index=False)
        .agg(预警次数=("预警类型", "count"), 最大故障时间=("故障时间", "max"))
        .sort_values(["预警级别", "预警次数", "最大故障时间"], ascending=[True, False, False])
        if not warning_df.empty
        else pd.DataFrame(columns=["生产线", "炉号", "预警级别", "预警次数", "最大故障时间"])
    )
    return warning_df, summary


def write_fault_warning_report(
    cycles: pd.DataFrame, furnace_list: list[str] | None = None
) -> tuple[Path, pd.DataFrame, pd.DataFrame]:
    print("\n正在生成故障预警...")
    prefix = scope_name(furnace_list)
    warnings, summary = detect_fault_warnings(cycles, furnace_list)
    output_path = OUTPUT_DIR / f"故障预警_{prefix}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_dataframe(writer, warnings, "故障预警明细")
        write_dataframe(writer, summary, "预警汇总")

    autosize_workbook(output_path)
    serious_count = int((warnings["预警级别"] == "严重").sum()) if not warnings.empty else 0
    print(f"  故障预警已输出：{output_path}，预警 {len(warnings)} 条，严重 {serious_count} 条")
    return output_path, warnings, summary


def run_fault_warning(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> Path:
    output_path, _, _ = write_fault_warning_report(cycles, furnace_list)
    return output_path


def run_daily(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[Path, Path]:
    workbook, daily_all, _ = write_daily_summary(cycles, furnace_list)
    prefix = scope_name(furnace_list)
    chart_name = "每日趋势图.png" if furnace_list is None else f"每日趋势图_{prefix}.png"
    chart = plot_summary_trend(daily_all, "日期", f"{prefix}每日生产趋势", OUTPUT_DIR / chart_name)
    return workbook, chart


def run_monthly(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> tuple[Path, Path]:
    workbook, monthly_all, _ = write_monthly_summary(cycles, furnace_list)
    prefix = scope_name(furnace_list)
    chart_name = "每月趋势图.png" if furnace_list is None else f"每月趋势图_{prefix}.png"
    chart = plot_summary_trend(monthly_all, "年月", f"{prefix}每月生产趋势", OUTPUT_DIR / chart_name)
    return workbook, chart


def run_all(cycles: pd.DataFrame, furnace_list: list[str] | None = None) -> AnalysisOutputs:
    print("\n>>> 开始运行全部分析")
    if furnace_list is None:
        furnace_count = cycles["炉号"].nunique()
        estimated = format_duration(furnace_count * 0.8)
        print(f"  --all 将为全区 {furnace_count} 个炉子生成单炉每日趋势图，预计额外耗时约 {estimated}。")
    furnace_workbook, furnace_chart = run_furnace_stats(cycles, furnace_list=furnace_list)
    daily_workbook, daily_chart = run_daily(cycles, furnace_list=furnace_list)
    monthly_workbook, monthly_chart = run_monthly(cycles, furnace_list=furnace_list)
    furnace_daily_workbook, furnace_daily_chart_dir, _ = run_furnace_daily_trends(cycles, furnace_list=furnace_list)
    anomaly_workbook = write_anomaly_report(cycles, furnace_list=furnace_list)
    print("\n全部分析完成")
    return AnalysisOutputs(
        furnace_workbook=furnace_workbook,
        furnace_chart=furnace_chart,
        daily_workbook=daily_workbook,
        daily_chart=daily_chart,
        monthly_workbook=monthly_workbook,
        monthly_chart=monthly_chart,
        furnace_daily_workbook=furnace_daily_workbook,
        furnace_daily_chart_dir=furnace_daily_chart_dir,
        anomaly_workbook=anomaly_workbook,
    )


def print_available_furnaces(cycles: pd.DataFrame) -> None:
    furnaces = sorted(cycles["炉号"].dropna().unique())
    print(f"\n可用炉号（{len(furnaces)} 个）：")
    for index in range(0, len(furnaces), 12):
        print("  " + ", ".join(furnaces[index : index + 12]))


def interactive_menu(cycles: pd.DataFrame) -> None:
    while True:
        print("\n" + "-" * 42)
        print("请选择分析功能：")
        print("  1. 炉子级统计 - 全区")
        print("  2. 炉子级统计 - 自选炉子")
        print("  3. 每日汇总 + 趋势图")
        print("  4. 每月汇总 + 趋势图")
        print("  5. 单炉每日趋势图 - 全区")
        print("  6. 单炉每日趋势图 - 自选炉子")
        print("  7. 异常检测报告")
        print("  8. 故障分析 + 热力图")
        print("  9. 故障预警")
        print("  10. 全部运行")
        print("  0. 退出")
        print("-" * 42)

        try:
            choice = input("请输入选项 (0-10)：").strip()
        except EOFError:
            print("\n已退出")
            return

        if choice == "1":
            run_furnace_stats(cycles)
        elif choice == "2":
            print_available_furnaces(cycles)
            user_input = input("\n请输入炉号或前缀，多个用逗号分隔：").strip()
            selected = resolve_furnaces(cycles, [user_input])
            if selected:
                run_furnace_stats(cycles, selected)
            else:
                print("未匹配到任何炉号")
        elif choice == "3":
            run_daily(cycles)
        elif choice == "4":
            run_monthly(cycles)
        elif choice == "5":
            run_furnace_daily_trends(cycles)
        elif choice == "6":
            print_available_furnaces(cycles)
            user_input = input("\n请输入炉号或前缀，多个用逗号分隔：").strip()
            selected = resolve_furnaces(cycles, [user_input])
            if selected:
                run_furnace_daily_trends(cycles, selected)
            else:
                print("未匹配到任何炉号")
        elif choice == "7":
            write_anomaly_report(cycles)
        elif choice == "8":
            run_fault_analysis(cycles)
        elif choice == "9":
            run_fault_warning(cycles)
        elif choice == "10":
            run_all(cycles)
        elif choice == "0":
            print("已退出")
            return
        else:
            print("无效选项，请重新输入")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="碳纳米管生产数据分析")
    parser.add_argument("--input", help="输入 Excel 文件路径；默认自动选择当前目录中的生产数据 .xlsx")
    parser.add_argument("--all", action="store_true", help="运行全部分析")
    parser.add_argument("--scope", nargs="*", help="筛选分析范围，可填炉号或前缀，例如 --scope E01,E02 或 --scope E")
    parser.add_argument("--furnace-stats", action="store_true", help="生成炉子级统计")
    parser.add_argument("--furnace", nargs="*", help="兼容旧参数：单独使用生成炉子级统计；与其他分析选项并用时作为筛选范围")
    parser.add_argument("--daily", action="store_true", help="生成每日汇总和趋势图")
    parser.add_argument("--monthly", action="store_true", help="生成每月汇总和趋势图")
    parser.add_argument("--furnace-daily-trend", action="store_true", help="生成每个炉子的单独每日趋势数据和趋势图")
    parser.add_argument("--anomaly", action="store_true", help="生成低产率异常检测报告")
    parser.add_argument("--fault-analysis", action="store_true", help="生成故障炉号排名和故障热力图")
    parser.add_argument("--fault-warning", action="store_true", help="生成故障预警报告")
    parser.add_argument("--list-furnaces", action="store_true", help="列出可用炉号")
    return parser.parse_args()


def collect_scope_selectors(args: argparse.Namespace) -> list[str] | None:
    selectors: list[str] = []
    if args.scope:
        selectors.extend(args.scope)
    if args.furnace is not None:
        selectors.extend(args.furnace)
    return selectors or None


def has_report_action(args: argparse.Namespace) -> bool:
    return any(
        [
            args.all,
            args.furnace_stats,
            args.daily,
            args.monthly,
            args.furnace_daily_trend,
            args.anomaly,
            args.fault_analysis,
            args.fault_warning,
        ]
    )


def main() -> None:
    args = parse_args()
    OUTPUT_DIR.mkdir(exist_ok=True)

    print("=" * 60)
    print("碳纳米管生产数据分析")
    print("=" * 60)

    input_file = find_input_file(args.input)
    cycles = load_and_clean_data(input_file)

    if args.list_furnaces:
        print_available_furnaces(cycles)
        return

    ran = False
    scope_selectors = collect_scope_selectors(args)
    selected_scope = resolve_furnaces(cycles, scope_selectors) if scope_selectors else None
    if args.all:
        run_all(cycles, furnace_list=selected_scope)
        ran = True
    else:
        legacy_furnace_only = args.furnace is not None and not has_report_action(args) and not args.scope
        if args.furnace_stats or legacy_furnace_only:
            run_furnace_stats(cycles, furnace_list=selected_scope)
            ran = True
        if args.daily:
            run_daily(cycles, furnace_list=selected_scope)
            ran = True
        if args.monthly:
            run_monthly(cycles, furnace_list=selected_scope)
            ran = True
        if args.furnace_daily_trend:
            run_furnace_daily_trends(cycles, furnace_list=selected_scope)
            ran = True
        if args.anomaly:
            write_anomaly_report(cycles, furnace_list=selected_scope)
            ran = True
        if args.fault_analysis:
            run_fault_analysis(cycles, furnace_list=selected_scope)
            ran = True
        if args.fault_warning:
            run_fault_warning(cycles, furnace_list=selected_scope)
            ran = True

    if not ran:
        if scope_selectors:
            print("已指定筛选范围，但未指定分析项；请同时添加 --daily、--monthly、--furnace-stats 等分析参数。")
            return
        interactive_menu(cycles)


if __name__ == "__main__":
    main()
